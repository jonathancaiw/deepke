import logging
from openpyxl import load_workbook
from global_config import *

file_handler = logging.FileHandler(LOG_FILENAME, encoding='utf-8')
stream_handler = logging.StreamHandler()
logging.basicConfig(format=LOG_FORMAT, datefmt=LOG_DATE_FORMAT, level=logging.INFO,
                    handlers=[file_handler, stream_handler])


def write_log(line, level=logging.INFO, newline=False):
    """
    写日志并输出到控制台
    :param line:
    :param level:
    :param newline:
    :return:
    """
    if newline:
        line = '\n' + str(line)

    if level == logging.CRITICAL:
        logging.critical(line)
    elif level == logging.ERROR:
        logging.error(line)
    elif level == logging.WARNING:
        logging.warning(line)
    elif level == logging.INFO:
        logging.info(line)
    elif level == logging.DEBUG:
        logging.debug(line)


def get_dict_from_xlsx(filename, columns, sheet=0):
    """
    解析xlsx，获取指定列的值列表
    :param filename:
    :param columns:
    :param sheet:
    :return:
    """
    key_values = {}

    workbook = load_workbook(filename)
    worksheet = workbook[workbook.sheetnames[sheet]]

    for column in worksheet.columns:
        column_name = column[0].value
        if column_name in columns:
            key_values[columns[column_name]] = []
            for i in range(1, len(column)):
                key_values[columns[column_name]].append(column[i].value)

    workbook.close()

    return key_values
